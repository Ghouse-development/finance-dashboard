"""
Excel Analyzer - Extracts complete structure and formatting data from .xlsx files.
Outputs a JSON file suitable for generating pixel-perfect HTML reproduction.

Usage: python excel_analyzer.py <excel_file_path> <output_dir>
"""

import sys
import os
import json
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Font


def color_to_hex(color):
    """Convert openpyxl color to CSS hex string."""
    if color is None:
        return None
    if color.type == "rgb" and color.rgb:
        rgb = str(color.rgb)
        if len(rgb) == 8:  # AARRGGBB
            return f"#{rgb[2:]}"
        elif len(rgb) == 6:
            return f"#{rgb}"
    if color.type == "theme":
        theme_colors = {
            0: "#FFFFFF", 1: "#000000", 2: "#E7E6E6", 3: "#44546A",
            4: "#4472C4", 5: "#ED7D31", 6: "#A5A5A5", 7: "#FFC000",
            8: "#5B9BD5", 9: "#70AD47"
        }
        return theme_colors.get(color.theme, None)
    if color.type == "indexed":
        indexed_colors = {
            0: "#000000", 1: "#FFFFFF", 2: "#FF0000", 3: "#00FF00",
            4: "#0000FF", 5: "#FFFF00", 6: "#FF00FF", 7: "#00FFFF",
            8: "#000000", 9: "#FFFFFF", 10: "#FF0000", 11: "#00FF00",
            12: "#0000FF", 13: "#FFFF00", 14: "#FF00FF", 15: "#00FFFF",
            16: "#800000", 17: "#008000", 18: "#000080", 19: "#808000",
            20: "#800080", 21: "#008080", 22: "#C0C0C0", 23: "#808080",
            64: "#000000"
        }
        return indexed_colors.get(color.indexed, None)
    return None


def border_side_to_dict(side):
    """Convert border side to dict."""
    if side is None or side.style is None:
        return None
    return {
        "style": side.style,
        "color": color_to_hex(side.color) if side.color else "#000000"
    }


def font_to_dict(font):
    """Convert font to dict."""
    if font is None:
        return {}
    return {
        "name": font.name,
        "size": font.size,
        "bold": font.bold,
        "italic": font.italic,
        "underline": font.underline,
        "strike": font.strikethrough,
        "color": color_to_hex(font.color) if font.color else None
    }


def alignment_to_dict(alignment):
    """Convert alignment to dict."""
    if alignment is None:
        return {}
    return {
        "horizontal": alignment.horizontal,
        "vertical": alignment.vertical,
        "wrap_text": alignment.wrap_text,
        "text_rotation": alignment.text_rotation,
        "indent": alignment.indent
    }


def fill_to_dict(fill):
    """Convert fill to dict."""
    if fill is None:
        return {}
    result = {"type": fill.fill_type}
    if fill.fgColor:
        result["fg_color"] = color_to_hex(fill.fgColor)
    if fill.bgColor:
        result["bg_color"] = color_to_hex(fill.bgColor)
    return result


def border_to_dict(border):
    """Convert border to dict."""
    if border is None:
        return {}
    result = {}
    for side_name in ["left", "right", "top", "bottom"]:
        side = getattr(border, side_name, None)
        side_dict = border_side_to_dict(side)
        if side_dict:
            result[side_name] = side_dict
    return result


def format_value(value, number_format):
    """Format a cell value according to its number format for display."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        try:
            if "yyyy/mm/dd" in number_format.lower() or "aaa" in number_format:
                weekdays_ja = ["月", "火", "水", "木", "金", "土", "日"]
                wd = weekdays_ja[value.weekday()]
                return f"{value.year}/{value.month:02d}/{value.day:02d}({wd})"
            elif "mmmm" in number_format.lower():
                return value.strftime("%Y/%m/%d")
            else:
                return value.strftime("%Y/%m/%d")
        except Exception:
            return str(value)
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)):
        nf = number_format or "General"
        if "%" in nf:
            decimal_places = 0
            if "." in nf:
                decimal_places = len(nf.split(".")[-1].replace("%", ""))
            return f"{value * 100:.{decimal_places}f}%"
        if "¥" in nf or "円" in nf:
            formatted = f"{int(round(value)):,}"
            if "¥" in nf:
                return f"¥{formatted}"
            if '円"' in nf or "円" in nf:
                return f"{formatted}円"
            return formatted
        if "#,##0" in nf:
            if "." in nf and "0" in nf.split(".")[-1]:
                after_dot = nf.split(".")[-1]
                decimal_places = 0
                for ch in after_dot:
                    if ch == "0" or ch == "#":
                        decimal_places += 1
                    else:
                        break
                return f"{value:,.{decimal_places}f}"
            if isinstance(value, float) and value == int(value):
                return f"{int(value):,}"
            return f"{int(round(value)):,}"
        if '"年"' in nf:
            return f"{int(value)}年"
        if '"ヶ"' in nf and '"月"' in nf:
            return f"{value:.1f}ヶ月"
        if "0.00" in nf:
            return f"{value:.2f}"
        if "0.0" in nf:
            return f"{value:.1f}"
        if "0.000" in nf:
            return f"{value:.3f}"
    return str(value)


def parse_print_area(print_area_str, sheet_title):
    """Parse print area string like 'Sheet1!$A$1:$DJ$105' into (min_col, min_row, max_col, max_row)."""
    if not print_area_str:
        return None
    area = str(print_area_str)
    if "!" in area:
        area = area.split("!")[-1]
    area = area.replace("$", "")
    if ":" not in area:
        return None
    start, end = area.split(":")
    import re
    m1 = re.match(r"([A-Z]+)(\d+)", start)
    m2 = re.match(r"([A-Z]+)(\d+)", end)
    if not m1 or not m2:
        return None
    return {
        "min_col": column_index_from_string(m1.group(1)),
        "min_row": int(m1.group(2)),
        "max_col": column_index_from_string(m2.group(1)),
        "max_row": int(m2.group(2))
    }


def analyze_sheet(ws, wb):
    """Analyze a single worksheet and return structured data."""
    sheet_data = {
        "name": ws.title,
        "dimensions": str(ws.dimensions),
        "min_row": ws.min_row,
        "max_row": ws.max_row,
        "min_col": ws.min_column,
        "max_col": ws.max_column,
    }

    # Page setup
    ps = ws.page_setup
    pm = ws.page_margins
    sheet_data["page_setup"] = {
        "paper_size": ps.paperSize,
        "orientation": ps.orientation,
        "scale": ps.scale,
        "fit_to_width": ps.fitToWidth,
        "fit_to_height": ps.fitToHeight,
    }
    sheet_data["page_margins"] = {
        "left": round(float(pm.left) * 25.4, 2) if pm.left else 0,
        "right": round(float(pm.right) * 25.4, 2) if pm.right else 0,
        "top": round(float(pm.top) * 25.4, 2) if pm.top else 0,
        "bottom": round(float(pm.bottom) * 25.4, 2) if pm.bottom else 0,
    }

    # Print area
    print_area = None
    if ws.print_area:
        pa = ws.print_area
        if isinstance(pa, list):
            pa = pa[0] if pa else None
        if pa:
            print_area = parse_print_area(str(pa), ws.title)
    sheet_data["print_area"] = print_area

    # Column widths
    col_widths = {}
    default_width = ws.sheet_format.defaultColWidth or 8.43
    sheet_data["default_col_width"] = default_width
    for col_letter, col_dim in ws.column_dimensions.items():
        if col_dim.width is not None:
            col_widths[col_letter] = {
                "width": col_dim.width,
                "hidden": col_dim.hidden or False
            }
    sheet_data["column_widths"] = col_widths

    # Row heights
    row_heights = {}
    default_height = ws.sheet_format.defaultRowHeight or 15.0
    sheet_data["default_row_height"] = default_height
    for row_idx, row_dim in ws.row_dimensions.items():
        if row_dim.height is not None:
            row_heights[str(row_idx)] = {
                "height": row_dim.height,
                "hidden": row_dim.hidden or False
            }
    sheet_data["row_heights"] = row_heights

    # Merged cells
    merged = []
    for merge_range in ws.merged_cells.ranges:
        merged.append({
            "range": str(merge_range),
            "min_col": merge_range.min_col,
            "min_row": merge_range.min_row,
            "max_col": merge_range.max_col,
            "max_row": merge_range.max_row
        })
    sheet_data["merged_cells"] = merged

    # Determine the area to scan
    if print_area:
        scan_min_row = print_area["min_row"]
        scan_max_row = print_area["max_row"]
        scan_min_col = print_area["min_col"]
        scan_max_col = print_area["max_col"]
    else:
        scan_min_row = ws.min_row or 1
        scan_max_row = ws.max_row or 1
        scan_min_col = ws.min_column or 1
        scan_max_col = ws.max_column or 1

    # Build merged cell lookup
    merged_lookup = {}
    for m in merged:
        for r in range(m["min_row"], m["max_row"] + 1):
            for c in range(m["min_col"], m["max_col"] + 1):
                if r == m["min_row"] and c == m["min_col"]:
                    merged_lookup[(r, c)] = {
                        "is_origin": True,
                        "colspan": m["max_col"] - m["min_col"] + 1,
                        "rowspan": m["max_row"] - m["min_row"] + 1
                    }
                else:
                    merged_lookup[(r, c)] = {"is_origin": False}

    # Cells with values and formatting
    cells = []
    for row_idx in range(scan_min_row, scan_max_row + 1):
        for col_idx in range(scan_min_col, scan_max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            merge_info = merged_lookup.get((row_idx, col_idx))

            # Skip non-origin cells in merged ranges
            if merge_info and not merge_info.get("is_origin"):
                continue

            cell_data = {
                "row": row_idx,
                "col": col_idx,
                "col_letter": get_column_letter(col_idx),
            }

            # Value
            val = cell.value
            nf = cell.number_format or "General"
            if val is not None:
                cell_data["display_value"] = format_value(val, nf)
                cell_data["number_format"] = nf
                if isinstance(val, (datetime, date)):
                    cell_data["raw_value"] = str(val)
                    cell_data["value_type"] = "date"
                elif isinstance(val, bool):
                    cell_data["raw_value"] = val
                    cell_data["value_type"] = "bool"
                elif isinstance(val, (int, float)):
                    cell_data["raw_value"] = val
                    cell_data["value_type"] = "number"
                elif isinstance(val, str):
                    if val.startswith("="):
                        cell_data["raw_value"] = val
                        cell_data["value_type"] = "formula"
                    else:
                        cell_data["raw_value"] = val
                        cell_data["value_type"] = "string"
                else:
                    cell_data["raw_value"] = str(val)
                    cell_data["value_type"] = "other"
            else:
                cell_data["display_value"] = ""
                cell_data["value_type"] = "empty"

            # Merge info
            if merge_info and merge_info.get("is_origin"):
                cell_data["colspan"] = merge_info["colspan"]
                cell_data["rowspan"] = merge_info["rowspan"]

            # Font
            cell_data["font"] = font_to_dict(cell.font)

            # Alignment
            cell_data["alignment"] = alignment_to_dict(cell.alignment)

            # Border
            cell_data["border"] = border_to_dict(cell.border)

            # Fill
            cell_data["fill"] = fill_to_dict(cell.fill)

            cells.append(cell_data)

    sheet_data["cells"] = cells
    return sheet_data


def analyze_workbook(file_path, output_dir):
    """Analyze an entire workbook and write JSON output."""
    print(f"Loading workbook: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=True)

    result = {
        "file": str(file_path),
        "sheet_names": wb.sheetnames,
        "sheets": []
    }

    for sheet_name in wb.sheetnames:
        print(f"  Analyzing sheet: {sheet_name}")
        ws = wb[sheet_name]
        sheet_data = analyze_sheet(ws, wb)
        result["sheets"].append(sheet_data)

    output_path = Path(output_dir) / "analysis.json"
    os.makedirs(output_dir, exist_ok=True)

    print(f"Writing analysis to: {output_path}")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2, default=str)

    print(f"Analysis complete. Output: {output_path}")
    print(f"  Sheets: {len(result['sheets'])}")
    for s in result["sheets"]:
        print(f"    {s['name']}: {len(s['cells'])} cells, {len(s['merged_cells'])} merges")

    return str(output_path)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python excel_analyzer.py <excel_file> <output_dir>")
        sys.exit(1)

    excel_file = sys.argv[1]
    output_dir = sys.argv[2]

    if not os.path.exists(excel_file):
        print(f"Error: File not found: {excel_file}")
        sys.exit(1)

    analyze_workbook(excel_file, output_dir)
